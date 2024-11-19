import pandas as pd
atten=pd.read_csv("/content/drive/MyDrive/Lab_09/input_attendance.csv")
print(atten)
atten['Date'] = atten['Timestamp'].apply(lambda x: x.split(' ')[0])
atten['Time'] = atten['Timestamp'].apply(lambda x: x.split(' ')[1])
print(atten)
with open("/content/drive/MyDrive/Lab_09/stud_list.txt") as file:
    students = [line.strip() for line in file]
classes_taken_dates = ["06/08/2024", "13/08/2024", "20/08/2024", "27/08/2024", "03/09/2024", "17/09/2024", "01/10/2024"]
classes_missed_dates = ["10/09/2024"]
exams_dates = ["24/09/2024"]
all_dates = classes_taken_dates + classes_missed_dates + exams_dates
from datetime import datetime
all_dates_sorted = sorted(all_dates, key=lambda date: datetime.strptime(date, "%d/%m/%Y"))

all_dates=all_dates_sorted


df=pd.DataFrame(0,index=students,columns=all_dates)
print(df)
start_time=datetime.strptime("18:00:00","%H:%M:%S").time()
end_time=datetime.strptime("20:00:00","%H:%M:%S").time()
for _,row in atten.iterrows():
  roll=row['Roll']
  date=row['Date']
  time=datetime.strptime(row['Time'],"%H:%M:%S").time()
  if roll in df.index and date in all_dates and start_time<=time<=end_time:
    df.at[roll,date]=df.at[roll,date]+1
print(df)
df['Total count of dates']=len(all_dates)
df['Total Attendance Marked']=df[all_dates].sum(axis=1)
df['Total Attendance allowed']=len(classes_taken_dates)*2
df['Proxy']=(df['Total Attendance Marked']-df['Total Attendance allowed'])
df.loc[df['Proxy']<0,'Proxy']=0
print(df)
output_file = "/content/drive/MyDrive/Lab_09/output.xlsx"
df.to_excel(output_file)
print(f"Excel file saved to {output_file}")
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


wb = load_workbook(output_file)
ws = wb.active


red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
no_fill = PatternFill()


for row in range(2, ws.max_row + 1):
    for col in range(2, len(classes_taken_dates) + 2):
        cell = ws.cell(row=row, column=col)
        value = cell.value


        if value == 2:
            cell.fill = green_fill
        elif value == 1:
            cell.fill = yellow_fill
        elif value >= 3:
            cell.fill = red_fill
        else:
            cell.fill = no_fill


wb.save(output_file)

print(f"Excel file with color formatting saved to {output_file}")
